import pandas
import re
from openpyxl import Workbook,load_workbook

def Regexp(string, pattern, flags=0):
    return re.search(pattern, string, flags)


def IsStrInclude(source, target, flags=0):
    if Regexp(target, source, flags):
        return True
    else:
        return False


class ExcelBasic(classmethod):
    def __init__(self, excel_path):
        self.excel_path = excel_path
        self.workbook = load_workbook(excel_path)
        self.current_sheet = self.workbook.active
        self.data = pandas.read_excel(excel_path, sheet_name=self.current_sheet.title, engine='openpyxl')
        # self.data = pandas.read_excel(excel_path, engine='openpyxl')

    def ReadData(self, *args):
        return self.data[args].values

    def WriteData(self, *args):
        pass

    def CreatSheet(self, sheet_name):
        pass

    def RemoveSheet(self, sheet_name):
        self.workbook.remove(sheet_name)

    def DropData(self, *args, axis=0):
        """
        :param args:
        :param axis: 0:删除行  1:删除列
        :return:
        """
        self.data.drop(*args, axis=axis)

    def DeleteData(self):
        pass

    def ChooseCurrentSheet(self, sheet_name):
        sheet_list = self.workbook.sheetnames
        if sheet_name in sheet_list:
            self.current_sheet = sheet_name
        else:
            print('{} not in sheet list'.format(sheet_name))

    def SaveWorkbook(self, excel_path=None):
        if not excel_path:
            excel_path = self.excel_path
        self.workbook.save(excel_path)
        # todo: 当前保存会导致其他表丢失
        pandas.DataFrame(self.data).to_excel(excel_path, sheet_name=self.current_sheet.title, index=False)

a= ExcelBasic(r'H:\PythonLearn\PythonCode\ExcelProcess\qualkitdo_sltest_trace.xlsx')
a.DropData([0])
print(a.current_sheet)
a.SaveWorkbook()
print(a.workbook.sheetnames)